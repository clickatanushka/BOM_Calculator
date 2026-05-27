# from flask import Flask, request, jsonify, render_template, send_file
# from flask_cors import CORS
# import os
# import io
# from quotation_engine import process_bom
# from ai_helper import ask_ai
# from price_engine import enrich_bom
# from smt_checker import check_smt
# from pdf_generator import generate_pdf
# from email_drafter import draft_quotation_email
# from agent import run_agent
# from pdf_bom_parser import parse_pdf_bom

# app = Flask(__name__)
# CORS(app)

# UPLOAD_FOLDER = "uploads"
# OUTPUT_FOLDER = "outputs"
# os.makedirs(UPLOAD_FOLDER, exist_ok=True)
# os.makedirs(OUTPUT_FOLDER, exist_ok=True)


# # ── HOME ──
# @app.route("/")
# def home():
#     return render_template("index.html")


# # ── UPLOAD BOM ──
# @app.route("/upload-bom", methods=["POST"])
# def upload_bom():
#     if "file" not in request.files:
#         return jsonify({"error": "No file uploaded"}), 400
#     file = request.files["file"]
#     if file.filename == "":
#         return jsonify({"error": "Empty filename"}), 400

#     filepath = os.path.join(UPLOAD_FOLDER, file.filename)
#     file.save(filepath)

#     try:
#         filename_lower = file.filename.lower()
#         if filename_lower.endswith(".pdf"):
#             print(f">>> PDF BOM detected: {file.filename}")
#             result = parse_pdf_bom(filepath)
#         elif filename_lower.endswith(".xlsx") or filename_lower.endswith(".xls"):
#             print(f">>> Excel BOM detected: {file.filename}")
#             result = process_bom(filepath)
#         else:
#             return jsonify({"error": "Unsupported file type. Use .xlsx, .xls, or .pdf"}), 400

#         result["filepath"] = filepath
#         return jsonify(result)

#     except Exception as e:
#         print("ERROR in upload_bom:", e)
#         return jsonify({"error": str(e)}), 500


# # ── ASK AI ──
# @app.route("/ask-ai", methods=["POST"])
# def ai_route():
#     data        = request.json
#     question    = data.get("question", "")
#     bom_summary = data.get("bom_summary", "")
#     if not question:
#         return jsonify({"error": "No question provided"}), 400
#     try:
#         answer = ask_ai(question, bom_summary)
#         return jsonify({"answer": answer})
#     except Exception as e:
#         return jsonify({"error": str(e)}), 500


# # ── ENRICH BOM ──
# # board_qty is passed from frontend.
# # price_engine calculates total_qty = component_qty × board_qty,
# # selects the correct tier price for that total_qty,
# # and stores:
# #   unit_price      = tier unit price
# #   per_board_cost  = unit_price × component_qty   (cost for 1 board)
# #   extended_price  = unit_price × total_qty        (total procurement cost)
# #
# # The route returns:
# #   total_cost      = SUM(extended_price) for all active lines
# #                     (total spend to buy all components for all boards)
# #   bom_cost_per_board = SUM(per_board_cost)
# #                        (component cost for a single board — used in quotation)
# @app.route("/enrich-bom", methods=["POST"])
# def enrich_bom_route():
#     data      = request.json
#     bom       = data.get("bom", [])
#     board_qty = int(data.get("board_qty", 1))

#     try:
#         enriched = enrich_bom(bom, board_qty=board_qty)

#         active = [r for r in enriched if r.get("dnp") != "Y"]

#         # Total procurement spend (all components × all boards)
#         total_cost = sum(
#             r.get("extended_price") or 0
#             for r in active
#             if r.get("extended_price") is not None
#         )

#         # Component cost per single board (for quotation page)
#         bom_cost_per_board = sum(
#             r.get("per_board_cost") or 0
#             for r in active
#             if r.get("per_board_cost") is not None
#         )

#         return jsonify({
#             "bom":               enriched,
#             "total_cost":        round(total_cost, 2),        # total procurement
#             "bom_cost_per_board": round(bom_cost_per_board, 4), # per-board BOM cost
#             "board_qty":         board_qty,
#         })

#     except Exception as e:
#         print("ERROR in enrich_bom_route:", e)
#         return jsonify({"error": str(e)}), 500


# # ── SMT FEASIBILITY CHECK ──
# @app.route("/check-smt", methods=["POST"])
# def check_smt_route():
#     data = request.json
#     bom  = data.get("bom", [])
#     try:
#         result = check_smt(bom)
#         return jsonify(result)
#     except Exception as e:
#         return jsonify({"error": str(e)}), 500


# # ── EXPORT PDF ──
# @app.route("/export-pdf", methods=["POST"])
# def export_pdf():
#     data = request.json
#     try:
#         buf = generate_pdf(data)
#         return send_file(
#             buf,
#             as_attachment=True,
#             download_name=f"{data.get('ref','quotation').replace('-','_')}.pdf",
#             mimetype="application/pdf"
#         )
#     except Exception as e:
#         return jsonify({"error": str(e)}), 500


# # ── EXPORT EXCEL ──
# @app.route("/export-quote", methods=["POST"])
# def export_quote():
#     from openpyxl import Workbook
#     from openpyxl.styles import Font, PatternFill, Alignment
#     from openpyxl.utils import get_column_letter

#     data               = request.json
#     bom                = data.get("bom", [])
#     customer           = data.get("customer", "Customer")
#     project            = data.get("project", "Project")
#     board_qty          = float(data.get("qty", 1))
#     asm_per_board      = float(data.get("asm", 0))
#     margin_pct         = float(data.get("margin", 0))
#     ref_no             = data.get("ref", "QT-001")

#     # Use per_board_cost — manual prices already stored on row by frontend
#     bom_cost_per_board = sum(
#         (r.get("per_board_cost") or r.get("unit_price") or 0)
#         for r in bom
#         if r.get("dnp") != "Y" and r.get("price_state") != "rfq"
#     )

#     rfq_parts = data.get("rfq_parts", [])

#     sub_per_board  = bom_cost_per_board + asm_per_board
#     markup_per_board = sub_per_board * (margin_pct / 100)
#     sell_per_board = sub_per_board + markup_per_board

#     wb = Workbook()
#     ws = wb.active
#     ws.title = "Quotation"

#     header_font = Font(bold=True, color="FFFFFF", size=11)
#     header_fill = PatternFill("solid", fgColor="1e2336")
#     green_font  = Font(bold=True, color="166534", size=12)
#     center      = Alignment(horizontal="center")

#     ws.merge_cells("A1:F1")
#     ws["A1"] = "EMS AI QUOTATION SYSTEM"
#     ws["A1"].font = Font(bold=True, size=14, color="1a56db")
#     ws["A1"].alignment = center

#     ws.merge_cells("A2:F2")
#     ws["A2"] = f"Ref: {ref_no}  |  Customer: {customer}  |  Project: {project}"
#     ws["A2"].alignment = center
#     ws.append([])

#     ws.append(["COST SUMMARY", "", f"Per Board (€)", f"Total ×{int(board_qty)} boards (€)"])
#     ws.append(["Component cost (BOM)",   "", f"{bom_cost_per_board:.4f}",  f"{bom_cost_per_board * board_qty:.2f}"])
#     ws.append(["Assembly cost",           "", f"{asm_per_board:.2f}",       f"{asm_per_board * board_qty:.2f}"])
#     ws.append(["Subtotal",                "", f"{sub_per_board:.4f}",       f"{sub_per_board * board_qty:.2f}"])
#     ws.append([f"Margin ({margin_pct:.0f}%)", "", f"{markup_per_board:.4f}", f"{markup_per_board * board_qty:.2f}"])
#     ws.append(["SELL PRICE",              "", f"{sell_per_board:.4f}",      f"{sell_per_board * board_qty:.2f}"])

#     sell_row = ws.max_row
#     for col in range(1, 5):
#         ws.cell(sell_row, col).font = green_font

#     ws.append([])
#     bom_headers = ["#", "Ref", "Description", "MPN", "Manufacturer",
#                    "Package", "Qty/Board", "Unit Price", "Per Board €", "Ext Total €", "Mount", "Status"]
#     ws.append(bom_headers)
#     hrow = ws.max_row
#     for col, _ in enumerate(bom_headers, 1):
#         cell = ws.cell(hrow, col)
#         cell.font = header_font
#         cell.fill = header_fill
#         cell.alignment = center

#     for i, row in enumerate(bom, 1):
#         unit      = row.get("unit_price") or 0
#         pb        = row.get("per_board_cost") or 0
#         ext       = row.get("extended_price") or 0
#         state     = row.get("price_state", "auto" if unit else "unpriced")
#         is_rfq    = state == "rfq"
#         is_manual = state == "manual"

#         status_str   = ("DNP" if row.get("dnp") == "Y" else
#                         "RFQ" if is_rfq else
#                         "MANUAL" if is_manual else
#                         "No Price" if not unit else "OK")
#         supplier_str = ("MANUAL" if is_manual else
#                         "RFQ" if is_rfq else
#                         row.get("nexar_supplier") or "—")

#         ws.append([
#             i,
#             row.get("ref", ""),
#             row.get("description", ""),
#             row.get("mpn", ""),
#             row.get("manufacturer", ""),
#             row.get("package", ""),
#             row.get("qty", 0),
#             f"€{unit:.4f}" if unit and not is_rfq else ("RFQ" if is_rfq else "—"),
#             f"€{pb:.4f}"   if pb   and not is_rfq else ("RFQ" if is_rfq else "—"),
#             f"€{ext:.2f}"  if ext  and not is_rfq else ("RFQ" if is_rfq else "—"),
#             supplier_str,
#             row.get("mount", ""),
#             status_str,
#         ])

#     # RFQ section
#     if rfq_parts:
#         ws.append([])
#         ws.append(["📋 PARTS REQUIRING MANUAL QUOTATION (RFQ)"])
#         ws.cell(ws.max_row, 1).font = Font(bold=True, color="F59E0B", size=11)
#         ws.append(["MPN", "Description", "Manufacturer", "Qty/Board", "Notes"])
#         hrow2 = ws.max_row
#         for col in range(1, 6):
#             ws.cell(hrow2, col).font = header_font
#             ws.cell(hrow2, col).fill = header_fill
#         for r in bom:
#             if r.get("price_state") == "rfq":
#                 ws.append([
#                     r.get("mpn", ""),
#                     r.get("description", ""),
#                     r.get("manufacturer", ""),
#                     r.get("qty", 0),
#                     "Awaiting supplier quotation",
#                 ])

#     col_widths = [4, 20, 35, 22, 20, 14, 8, 12, 12, 12, 14, 8, 10]
#     for i, w in enumerate(col_widths, 1):
#         ws.column_dimensions[get_column_letter(i)].width = w

#     buf = io.BytesIO()
#     wb.save(buf)
#     buf.seek(0)
#     return send_file(buf, as_attachment=True,
#                      download_name=f"{ref_no.replace('-', '_')}.xlsx",
#                      mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# # ── DRAFT EMAIL ──
# @app.route("/draft-email", methods=["POST"])
# def draft_email_route():
#     data = request.json
#     try:
#         email_text = draft_quotation_email(data)
#         return jsonify({"email": email_text})
#     except Exception as e:
#         return jsonify({"error": str(e)}), 500


# # ── RUN FULL AGENT ──
# @app.route("/run-agent", methods=["POST"])
# def run_agent_route():
#     data     = request.json
#     filepath = data.get("filepath", "")

#     if not filepath or not os.path.exists(filepath):
#         return jsonify({"error": "File not found. Upload BOM first."}), 400

#     try:
#         result = run_agent(
#             filepath=filepath,
#             customer=data.get("customer", "Customer"),
#             project =data.get("project",  "Project"),
#             qty     =float(data.get("qty",      100)),
#             asm_cost=float(data.get("asm_cost", 8.5)),
#             margin  =float(data.get("margin",   20)),
#             ref     =data.get("ref", "QT-001"),
#         )
#         return jsonify(result)
#     except Exception as e:
#         return jsonify({"error": str(e)}), 500


# if __name__ == "__main__":
#     app.run(debug=True, port=5000)

from flask import Flask, request, jsonify, render_template, send_file
from flask_cors import CORS
import os
import io
from quotation_engine import process_bom
from ai_helper import ask_ai
from price_engine import enrich_bom
from smt_checker import check_smt
from pdf_generator import generate_pdf
from email_drafter import draft_quotation_email
from agent import run_agent
from pdf_bom_parser import parse_pdf_bom

app = Flask(__name__)
CORS(app)

UPLOAD_FOLDER = "uploads"
OUTPUT_FOLDER = "outputs"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)


# ── HOME ──
@app.route("/")
def home():
    return render_template("index.html")


# ── UPLOAD BOM ──
@app.route("/upload-bom", methods=["POST"])
def upload_bom():
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    file = request.files["file"]
    if file.filename == "":
        return jsonify({"error": "Empty filename"}), 400

    filepath = os.path.join(UPLOAD_FOLDER, file.filename)
    file.save(filepath)

    try:
        filename_lower = file.filename.lower()
        if filename_lower.endswith(".pdf"):
            print(f">>> PDF BOM detected: {file.filename}")
            result = parse_pdf_bom(filepath)
        elif filename_lower.endswith(".xlsx") or filename_lower.endswith(".xls"):
            print(f">>> Excel BOM detected: {file.filename}")
            result = process_bom(filepath)
        else:
            return jsonify({"error": "Unsupported file type. Use .xlsx, .xls, or .pdf"}), 400

        result["filepath"] = filepath
        return jsonify(result)

    except Exception as e:
        print("ERROR in upload_bom:", e)
        return jsonify({"error": str(e)}), 500


# ── ASK AI ──
@app.route("/ask-ai", methods=["POST"])
def ai_route():
    data        = request.json
    question    = data.get("question", "")
    bom_summary = data.get("bom_summary", "")
    if not question:
        return jsonify({"error": "No question provided"}), 400
    try:
        answer = ask_ai(question, bom_summary)
        return jsonify({"answer": answer})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── ENRICH BOM ──
# board_qty is passed from frontend.
# price_engine calculates total_qty = component_qty × board_qty,
# selects the correct tier price for that total_qty,
# and stores:
#   unit_price      = tier unit price
#   per_board_cost  = unit_price × component_qty   (cost for 1 board)
#   extended_price  = unit_price × total_qty        (total procurement cost)
#
# The route returns:
#   total_cost      = SUM(extended_price) for all active lines
#                     (total spend to buy all components for all boards)
#   bom_cost_per_board = SUM(per_board_cost)
#                        (component cost for a single board — used in quotation)
@app.route("/enrich-bom", methods=["POST"])
def enrich_bom_route():
    data      = request.json
    bom       = data.get("bom", [])
    board_qty = int(data.get("board_qty", 1))

    try:
        enriched = enrich_bom(bom, board_qty=board_qty)

        active = [r for r in enriched if r.get("dnp") != "Y"]

        # Total procurement spend (all components × all boards)
        total_cost = sum(
            r.get("extended_price") or 0
            for r in active
            if r.get("extended_price") is not None
        )

        # Component cost per single board (for quotation page)
        bom_cost_per_board = sum(
            r.get("per_board_cost") or 0
            for r in active
            if r.get("per_board_cost") is not None
        )

        return jsonify({
            "bom":               enriched,
            "total_cost":        round(total_cost, 2),        # total procurement
            "bom_cost_per_board": round(bom_cost_per_board, 4), # per-board BOM cost
            "board_qty":         board_qty,
        })

    except Exception as e:
        print("ERROR in enrich_bom_route:", e)
        return jsonify({"error": str(e)}), 500


# ── SMT FEASIBILITY CHECK ──
@app.route("/check-smt", methods=["POST"])
def check_smt_route():
    data = request.json
    bom  = data.get("bom", [])
    try:
        result = check_smt(bom)
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── EXPORT PDF ──
@app.route("/export-pdf", methods=["POST"])
def export_pdf():
    data = request.json
    try:
        buf = generate_pdf(data)
        return send_file(
            buf,
            as_attachment=True,
            download_name=f"{data.get('ref','quotation').replace('-','_')}.pdf",
            mimetype="application/pdf"
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── EXPORT EXCEL ──
@app.route("/export-quote", methods=["POST"])
def export_quote():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    data               = request.json
    bom                = data.get("bom", [])
    customer           = data.get("customer", "Customer")
    project            = data.get("project", "Project")
    board_qty          = float(data.get("qty", 1))
    asm_per_board      = float(data.get("asm", 0))
    margin_pct         = float(data.get("margin", 0))
    ref_no             = data.get("ref", "QT-001")

    # Use per_board_cost — manual prices already stored on row by frontend
    bom_cost_per_board = sum(
        (r.get("per_board_cost") or r.get("unit_price") or 0)
        for r in bom
        if r.get("dnp") != "Y" and r.get("price_state") != "rfq"
    )

    rfq_parts = data.get("rfq_parts", [])

    sub_per_board  = bom_cost_per_board + asm_per_board
    markup_per_board = sub_per_board * (margin_pct / 100)
    sell_per_board = sub_per_board + markup_per_board

    wb = Workbook()
    ws = wb.active
    ws.title = "Quotation"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1e2336")
    green_font  = Font(bold=True, color="166534", size=12)
    center      = Alignment(horizontal="center")

    ws.merge_cells("A1:F1")
    ws["A1"] = "EMS AI QUOTATION SYSTEM"
    ws["A1"].font = Font(bold=True, size=14, color="1a56db")
    ws["A1"].alignment = center

    ws.merge_cells("A2:F2")
    ws["A2"] = f"Ref: {ref_no}  |  Customer: {customer}  |  Project: {project}"
    ws["A2"].alignment = center
    ws.append([])

    ws.append(["COST SUMMARY", "", f"Per Board (€)", f"Total ×{int(board_qty)} boards (€)"])
    ws.append(["Component cost (BOM)",   "", f"{bom_cost_per_board:.4f}",  f"{bom_cost_per_board * board_qty:.2f}"])
    ws.append(["Assembly cost",           "", f"{asm_per_board:.2f}",       f"{asm_per_board * board_qty:.2f}"])
    ws.append(["Subtotal",                "", f"{sub_per_board:.4f}",       f"{sub_per_board * board_qty:.2f}"])
    ws.append([f"Margin ({margin_pct:.0f}%)", "", f"{markup_per_board:.4f}", f"{markup_per_board * board_qty:.2f}"])
    ws.append(["SELL PRICE",              "", f"{sell_per_board:.4f}",      f"{sell_per_board * board_qty:.2f}"])

    sell_row = ws.max_row
    for col in range(1, 5):
        ws.cell(sell_row, col).font = green_font

    ws.append([])
    bom_headers = ["#", "Ref", "Description", "MPN", "Manufacturer",
                   "Package", "Qty/Board", "Unit Price", "Per Board €", "Ext Total €", "Mount", "Status"]
    ws.append(bom_headers)
    hrow = ws.max_row
    for col, _ in enumerate(bom_headers, 1):
        cell = ws.cell(hrow, col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    for i, row in enumerate(bom, 1):
        unit      = row.get("unit_price") or 0
        pb        = row.get("per_board_cost") or 0
        ext       = row.get("extended_price") or 0
        state     = row.get("price_state", "auto" if unit else "unpriced")
        is_rfq    = state == "rfq"
        is_manual = state == "manual"

        status_str   = ("DNP" if row.get("dnp") == "Y" else
                        "RFQ" if is_rfq else
                        "MANUAL" if is_manual else
                        "No Price" if not unit else "OK")
        supplier_str = ("MANUAL" if is_manual else
                        "RFQ" if is_rfq else
                        row.get("nexar_supplier") or "—")

        ws.append([
            i,
            row.get("ref", ""),
            row.get("description", ""),
            row.get("mpn", ""),
            row.get("manufacturer", ""),
            row.get("package", ""),
            row.get("qty", 0),
            f"€{unit:.4f}" if unit and not is_rfq else ("RFQ" if is_rfq else "—"),
            f"€{pb:.4f}"   if pb   and not is_rfq else ("RFQ" if is_rfq else "—"),
            f"€{ext:.2f}"  if ext  and not is_rfq else ("RFQ" if is_rfq else "—"),
            supplier_str,
            row.get("mount", ""),
            status_str,
        ])

    # RFQ section
    if rfq_parts:
        ws.append([])
        ws.append(["📋 PARTS REQUIRING MANUAL QUOTATION (RFQ)"])
        ws.cell(ws.max_row, 1).font = Font(bold=True, color="F59E0B", size=11)
        ws.append(["MPN", "Description", "Manufacturer", "Qty/Board", "Notes"])
        hrow2 = ws.max_row
        for col in range(1, 6):
            ws.cell(hrow2, col).font = header_font
            ws.cell(hrow2, col).fill = header_fill
        for r in bom:
            if r.get("price_state") == "rfq":
                ws.append([
                    r.get("mpn", ""),
                    r.get("description", ""),
                    r.get("manufacturer", ""),
                    r.get("qty", 0),
                    "Awaiting supplier quotation",
                ])

    col_widths = [4, 20, 35, 22, 20, 14, 8, 12, 12, 12, 14, 8, 10]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f"{ref_no.replace('-', '_')}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ── DRAFT EMAIL ──
@app.route("/draft-email", methods=["POST"])
def draft_email_route():
    data = request.json
    try:
        email_text = draft_quotation_email(data)
        return jsonify({"email": email_text})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── RUN FULL AGENT ──
@app.route("/run-agent", methods=["POST"])
def run_agent_route():
    data     = request.json
    filepath = data.get("filepath", "")

    if not filepath or not os.path.exists(filepath):
        return jsonify({"error": "File not found. Upload BOM first."}), 400

    try:
        result = run_agent(
            filepath=filepath,
            customer=data.get("customer", "Customer"),
            project =data.get("project",  "Project"),
            qty     =float(data.get("qty",      100)),
            asm_cost=float(data.get("asm_cost", 8.5)),
            margin  =float(data.get("margin",   20)),
            ref     =data.get("ref", "QT-001"),
        )
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


if __name__ == "__main__":
    app.run(debug=True, port=5000)